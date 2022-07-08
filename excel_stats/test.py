

def excel_file():
    from openpyxl import load_workbook
    wb = load_workbook(filename=input(
        "Please enter the excel file directory!\n"))
    sheet = wb.active
    for names in sheet.iter_rows(min_row=1,
                                 max_row=1,
                                 values_only=True):
        col_names = names
    data = [[sheet.cell(row=i, column=j).value for j in range(
            1, sheet.max_column+1)] for i in range(2, sheet.max_row+1)]
    data_dict = {}
    for i, col_name in enumerate(col_names):
        data_dict[col_name] = [row[i] for row in data]
    return (col_names, data_dict)


list1, dict1 = excel_file()
print("\nThe excel file contains the following columns:\n", list1,
      "\nThis is the whole excel file in a dictionary:\n", dict1)
