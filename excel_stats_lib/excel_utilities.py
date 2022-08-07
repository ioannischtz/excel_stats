from openpyxl import load_workbook
# File path to test the code: E:\DEV\excel_stats\data\soundnames-LTM coordinates.xlsx
inputfile = "E:\DEV\excel_stats\data\pytestfile.xlsx"


def getExcelSheet(inputfile):
    wb = load_workbook(filename=inputfile)
    sheet = wb.active

    return sheet


def getSheetNames(sheet, b_rows=False):
    if b_rows:
        for names in sheet.iter_cols(min_col=1,
                                     max_col=1,
                                     values_only=True):
            sheet_names = names
    else:
        for names in sheet.iter_rows(min_row=1,
                                     max_row=1,
                                     values_only=True):
            sheet_names = names
    return sheet_names


def getDataList(sheet, b_rows=False):
    if b_rows:
        data_list = [[sheet.cell(row=i, column=j).value for j in range(
            1, sheet.max_column+1)] for i in range(2, sheet.max_row+1)]
    else:
        data_list = [[sheet.cell(row=i, column=j).value for i in range(
            2, sheet.max_row+1)] for j in range(1, sheet.max_column+1)]
    return data_list


def getDataDict(sheet, col_names, data_list_rows):
    data_dict = {}
    for i, sheet_name in enumerate(col_names):
        data_dict[sheet_name] = [row[i] for row in data_list_rows]

    return data_dict

# sheet = getExcelSheet(inputfile)
# col_names = getSheetNames(sheet)
# data_list_rows = getDataList(sheet, b_rows=True)
# print(getDataDict(sheet, col_names, data_list_rows))