# ---- Open excel sheet ----
from openpyxl import load_workbook
# File path to test the code: E:\DEV\excel_stats\data\soundnames-LTM coordinates.xlsx
inputfile = "E:\DEV\excel_stats\data\soundnames-LTM coordinates.xlsx"


def extractExceldata(inputfile):
    wb = load_workbook(filename=inputfile)
    sheet = wb.active
    for names in sheet.iter_rows(min_row=1,
                                 max_row=1,
                                 values_only=True):
        col_names = names

    data_list_rows = [[sheet.cell(row=i, column=j).value for j in range(
        1, sheet.max_column+1)] for i in range(2, sheet.max_row+1)]

    data_list_columns = [[sheet.cell(row=i, column=j).value for i in range(
        2, sheet.max_row+1)] for j in range(1, sheet.max_column+1)]

    data_dict = {}
    for i, col_name in enumerate(col_names):
        data_dict[col_name] = [row[i] for row in data_list_rows]

    return (col_names, data_dict, data_list_rows, data_list_columns)


# 1 Location
# 1.1 Arithmetic Mean


def arithMean_col(columnName, col_names, data_list_columns):
    col_name = col_names.index(columnName)
    max_r = len(data_list_columns[col_name])
    mean = sum(data_list_columns[col_name])/(max_r-1)

    return mean

# 1.2 median


def median(columnName, col_names, data_list_columns):
    col_name = col_names.index(columnName)
    sortedList = data_list_columns[col_name]
    sortedList.sort()
    m = len(sortedList)
    if m % 2 == 0:
        med = (sortedList[int(m/2-1)]+sortedList[int(m/2)])/2
    elif m % 2 == 1:
        med = (sortedList[int(m/2)])

    return med


# 1.3 mode
def unique(list1):

    # initialize a null list
    unique_list = []

    # traverse for all elements
    for x in list1:
        # check if exists in unique_list or not
        if x not in unique_list:
            unique_list.append(x)
    return unique_list


def stat_mode(columnName, col_names, data_list_columns):
    col_name = col_names.index(columnName)
    list1 = data_list_columns[col_name]
    unique_list = unique(list1)
    unXindex = [0 for x in range(0, len(unique_list))]
    for x in list1:
        for index, ux in enumerate(unique_list):
            if x == ux:
                unXindex[index] += 1
    mode_index = max(unXindex)
    mode = unique_list[mode_index]
    return mode
