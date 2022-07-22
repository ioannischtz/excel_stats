import matplotlib.pyplot as plt
import main_lib
from openpyxl import load_workbook

filename = "E:\DEV\excel_stats\data\soundnames-LTM coordinates.xlsx"
col_names, data_dict, data_list_rows, data_list_columns = main_lib.extractExceldata(
    filename)
wb = load_workbook(
    filename="E:\DEV\excel_stats\data\soundnames-LTM coordinates.xlsx")
sheet = wb.active

new_list = [[sheet.cell(row=i, column=j).value for j in range(
    3, sheet.max_column+1)] for i in range(2, sheet.max_row+1)]

x = []
for i, name in enumerate(col_names[2:]):
    x.append(i+1)

titles = data_list_columns[1]

fig, ax = plt.subplots()
ax.bar(x, new_list[0], width=1, edgecolor="red", linewidth=1)
ax.set(xlim=(0, 5), ylim=(0, 10),
       title=titles[0], xticks=x, xticklabels=col_names[2:])


plt.show()
